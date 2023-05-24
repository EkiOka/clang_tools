
# import standard library
import uuid
import datetime
import hashlib
import os
import sys
import pathlib
import shutil
import glob
import json
import codecs
import tkinter
import logging
import logging.config
import traceback
import re
import string
import inspect
import platform
import xml.etree.ElementTree as ET

# import pip install library                                  python -m pip install --upgrade pip
import jinja2                                               # pip install jinja2
import openpyxl                                             # pip install openpyxl
import markdown                                             # pip install markdown
from markdown_include.include import MarkdownInclude        # pip install markdown-include
from markdown_checklist.extension import ChecklistExtension # pip install markdown-checklist
from md2html_links import CustomLinkExtension               # download https://github.com/ricmua/markdown-md2html_links
                                                            # and copy md2html_links.py to python lib
import yaml                                                 # pip install pyyaml

class ul:
    """use case library
    """

    #------------------------------------------------------------------------
    # COMMON
    #------------------------------------------------------------------------
    id:str = "a48c5c3ad4e94017bcc275492c101193"
    default_encoding:str="utf-8"
    separate_path_dir  = os.sep
    separate_path_ext  = os.extsep
    separate_path_list = os.pathsep
    #------------------------------------------------------------------------
    # DEBUG
    #------------------------------------------------------------------------
    @staticmethod
    def is_vscode_debug():
        res = False
        if "debugpy" in sys.modules:
            res = True
        return res
    @staticmethod
    def log_enable(cfg=None):
        if cfg == None:            
            cfg = {
                "version": 1,
                "disable_existing_loggers": False,
                "formatters": {
                    # フォーマットの説明は以下参照
                    # https://docs.python.org/ja/3/library/logging.html#logrecord-attributes
                    "standard": {
                        "format": "python(%(levelname)s) > %(message)s"
                    },
                    "add_position": {
                        "format": "python(%(levelname)s) > %(funcName)s %(lineno)s %(message)s"
                    }
                },

                "handlers": {
                    "console": {
                        "class": "logging.StreamHandler",
                        "level": "DEBUG",
                        "formatter": "standard",
                        "stream": "ext://sys.stdout"
                    },
                    # "file": {
                    #     "class": "logging.FileHandler",
                    #     "level": "INFO",
                    #     "formatter": "standard",
                    #     "filename": f"{__name__}.log"
                    # },
                    "null": {
                        "class": "logging.NullHandler"
                    }
                },

                "loggers": {
                    f"{__name__}": {
                        "level": "DEBUG",
                        "handlers": ["console"],
                        "propagate": False
                    }
                },

                "root": {
                    "level": "INFO"
                }
            }

        logging.config.dictConfig(cfg)
    @staticmethod
    def log_debug(text:str,name:str=__name__):
        log = logging.getLogger(name)
        log.debug(text)
    @staticmethod
    def log_info(text:str,name:str=__name__):
        log = logging.getLogger(name)
        log.info(text)
    @staticmethod
    def log_warning(text:str,name:str=__name__):
        log = logging.getLogger(name)
        log.info(text)
    @staticmethod
    def log_critical(text:str,name:str=__name__):
        log = logging.getLogger(name)
        log.critical(text)
    @staticmethod
    def get_caller_file_name():
        return inspect.currentframe().f_back.f_back.f_code.co_filename
    @staticmethod
    def get_caller_function_name():
        return inspect.currentframe().f_back.f_back.f_code.co_name
    @staticmethod
    def get_caller_line_no():
        return inspect.currentframe().f_back.f_back.f_lineno
    @staticmethod
    def get_cur_file_name():
        return inspect.currentframe().f_back.f_code.co_filename
    @staticmethod
    def get_cur_function_name():
        return inspect.currentframe().f_back.f_code.co_name
    @staticmethod
    def get_cur_line_no():
        return inspect.currentframe().f_back.f_lineno
    #========================================================================
    #
    # DATA
    #
    #========================================================================
    @staticmethod
    def loop(data:list,method,last_method):
        """ループ処理を行う

        最後の要素だけセパレータを付けない場合など、
        少し処理が違う場合などに使用すること
        """
        res = list()
        # 途中の要素を処理
        for d in data[:-1]:
            res.append(method(d))
        res.append(last_method(data[-1]))
        return res
    #------------------------------------------------------------------------
    # TEXT
    #------------------------------------------------------------------------
    @staticmethod
    def get_text_tail(text):
        res = ""
        if len(text) > 0:
            res = text[len(text) - 1]
        return res
    @staticmethod
    def get_uuid()->str:
        return uuid.uuid4().hex.upper()
    @staticmethod
    def gen_uuid_upper()->str:
        return uuid.uuid4().hex.upper()
    @staticmethod
    def gen_uuid_lower()->str:
        return uuid.uuid4().hex.lower()
    #========================================================================
    #
    # FILE CONTROL
    #
    #========================================================================

    #------------------------------------------------------------------------
    # FILE/DIRECTORY/PATH
    #------------------------------------------------------------------------
    @staticmethod
    def get_cur_dir()->str:
        """カレントディレクトリを取得する
        """
        return os.getcwd().replace("\\","/")
    @staticmethod
    def set_cur_dir(path:str):
        """カレントディレクトリを変更する
        """
        os.chdir(path)
    @staticmethod
    def cnv_res_to_abs_path(path:str):
        p = pathlib.Path(path)
        res = p.resolve()
        return res
    @staticmethod
    def cnv_no_ext_path(path:str):
        res = os.path.splitext(path)[0]
        return res
    @staticmethod
    def cnv_file_name(path:str):
        res = os.path.basename(path)
        return res
    @staticmethod
    def cnv_file_name_none_ext(path:str)->str:
        res = os.path.splitext(os.path.basename(path))[0]
        return res
    @staticmethod
    def cnv_dir_path(path:str)->str:
        res = os.path.dirname(path)
        return res
    @staticmethod
    def get_parent_dir(src:str)->str:
        """親ディレクトリ取得

        Parameters
        ----------
        src : str
            親ディレクトリを取得する元となるディレクトリ

        Returns
        -------
        str
            親ディレクトリのパス
        """
        p = pathlib.Path(src)
        res = str(p.parent)
        return res
    @staticmethod
    def get_file_name_ext(path:str)->str:
        res = os.path.basename(path)
        return res
    @staticmethod
    def get_dict_path(path:str, path_type:str = "auto" )->dict:
        """パスを様々な形式に変換します

        パス関係は複雑にできており、色々と手間が必要なため
        面倒な場合は本関数にパスを入れて欲しい結果のみを使用すること

        Parameters
        ----------
        path : str
            convert target.\n
            変換対象となるパスです。
        path_type : str, optional
            target type, by default "auto".\n
            You can set "file", "dir" and "auto" to parameter.\n
            変換対象となるパスの種別です。\n
            "file","dir","auto"を設定できます。\n
            設定しない場合は"auto"となります。

        Returns
        -------
        dict
            辞書型に変換したパスを格納します。\n
            \n
            abs_dir/         : 絶対パスのディレクトリのみ                      \n
            abs_dir/file     : 絶対パスのディレクトリ+ファイル名(拡張子なし)   \n
            abs_dir/file.ext : 絶対パスのディレクトリ+ファイル名(拡張子あり)   \n
            rel_dir/         : 相対パスのディレクトリのみ                      \n
            rel_dir/file     : 相対パスのディレクトリ+ファイル名(拡張子なし)   \n
            rel_dir/file.ext : 相対パスのディレクトリ+ファイル名(拡張子なし)   \n
            .ext             : 拡張子(dot付き)                                 \n
            ext              : 拡張子(dotなし)                                 \n
            extension_all    : 拡張子(*.tar.gzのようなパスの場合。先頭の.以降) \n
            extension        : 拡張子(*.tar.gzのようなパスの場合。最後の.以降) \n
            file             : ファイル名(拡張子なし)(最後のdotまで)           \n
            file_name        : ファイル名(拡張子なし)(最後のdotまで)           \n
            file.ext         : ファイル名(拡張子あり)                          \n
            file_label       : ファイル名(拡張子なし)(最初のdotまで)           \n

        Raises
        ------
        Exception
            path_typeに"auto"を設定している場合のみに発生します。\n
            "dir/makefile"のように拡張子の付いていない場合で、
            実際存在していないファイルパスを扱っている場合、
            自動で判別できないため発生する例外となります。\n
            解消するには引数のpath_typeに"file"か"dir"を設定して下さい。
        Exception
            ファイル名のセパレータによる分割に失敗したと判断しました。\n
            念のためのガード処理のため原因は不明となります。
        Exception
            パスを生成する段階においてpath_typeが\n
            確定していない場合に発生します。\n
            本来ありえない状況なので関数のリファクタリングに\n
            失敗している可能性があります。

        Notes
        -----
        ただし、大きなプロジェクトに使用するには
        処理が重いため注意すること
        """
        # 結果を辞書型に入れてソートをかけて返す
        res = dict()

        is_abs   = os.path.isabs(path)
        is_exist = os.path.exists(path)
        is_file  = os.path.isfile(path)
        is_dir   = os.path.isdir(path)

        # 一旦全部形式を/区切りの絶対パスに揃える
        slash_src_path = path.replace("\\","/")
        slash_abs_path = ""

        if is_abs:
            slash_abs_path = slash_src_path
        else:
            slash_abs_path = os.path.abspath(slash_src_path)

        if path_type == "auto":
            # パスの明確な指示がない場合
            # 自動判定を行う

            # 一旦文字列を空にし
            # 出口にてロジックに問題がないか確認する
            path_type = ""

            # 実際に存在するパスの場合
            # 確認結果をそのまま反映
            if is_exist:
                if is_file:
                    path_type = "file"
                else:
                    path_type = "dir"
            else:
                # 実際に存在しないパスの場合
                # 1. 末尾に / があればディレクトリ
                # 2. 無い場合は拡張子を確認
                #    ただし、ディレクトリにも.は使えるため
                #    既知のディレクトリかどうかで判断する
                tail = ul.get_text_tail(slash_abs_path)

                if tail == "/":
                    path_type = "dir"
                else:
                    # ファイル名として取得
                    tmp_file_name_ext = os.path.basename(slash_abs_path)

                    tmp_extension     = os.path.splitext(
                        os.path.basename(path))[1]

                    has_dot = "." in tmp_file_name_ext

                    if has_dot:
                        # .入りパスは基本ファイルと判定
                        # ディレクトリパターンに
                        # 一致したらディレクトリと判定
                        dir_mat_name = tmp_file_name_ext in [
                            ".vscode",
                            ".svn",
                            ".git"
                            ".aws"
                            ]
                        dir_mat_ext = tmp_extension in [
                            ".files"
                        ]

                        path_type = "file"
                        if dir_mat_name:
                            path_type = "dir"
                        elif dir_mat_ext:
                            path_type = "dir"
                    else:
                        # .なしのパスは判定が難しい。
                        # makefileはファイルに思えるが、
                        # *.mkファイルを入れるフォルダに
                        # makefileと付ける場合もあるので。
                        # よって、下記の例外が発生する場合
                        # path_typeを設定して使用してください。
                        raise Exception(
                            "パスのファイル判定ができませんでした。\
                            cnv関数にpath_type引数を設定して使用してください。")

        # 上記までの処理で下記の結果は流用可能
        # ・is_abs
        # ・is_exist
        # ・is_file
        # ・is_dir
        # ・slash_src_path
        # ・slash_abs_path
        # そのため関数の分離は行わない。

        # まずは / のまま各パスを計算
        slash_abs_dir = ""
        slash_rel_dir = ""
        file_name     = ""
        file_name_ext = ""
        extension     = "" # 最後の.から後ろを拡張子とした場合
        extension_all = "" # 最初の.から後ろを全て拡張子とした場合

        # パスがファイルである場合と
        # ディレクトリである場合で
        # 結果が変わるため処理を分離
        match path_type:
            case "file":
                slash_abs_dir  = os.path.dirname(slash_abs_path)
                file_name_ext = os.path.basename(slash_abs_path)

                tmp = file_name_ext.split(os.extsep)
                tmp_len = len(tmp)

                if tmp_len<=0:
                    raise Exception(
                        f"'{slash_abs_path}'のファイル名分離に失敗しました。\
                        {file_name_ext}を分割しようとしましたが中身が空です。")
                elif tmp_len == 1:
                    # 拡張子なしファイル名
                    file_name     = tmp[0]
                    file_label    = tmp[0]
                    # extension     = ""
                    # extension_all = ""
                elif tmp_len == 2:
                    # 通常ファイル名(*.*)
                    file_name     = tmp[0]
                    file_label    = tmp[0]
                    extension     = tmp[1]
                    extension_all = tmp[1]
                else:
                    # セパレータが複数ある場合
                    file_name     = tmp[0]
                    file_label    = tmp[0]
                    for t in tmp[1:-1]:
                        file_name     = file_name + os.extsep + t
                        extension_all = extension_all + os.extsep + t
                    extension     = os.extsep + tmp[-1]
                    extension_all = extension_all + os.extsep + tmp[-1]

                try:
                    slash_rel_dir = os.path.relpath(slash_abs_dir)
                except ValueError:
                    # 指定パスとカレントパスがドライブ違いの場合
                    # 絶対パスとして扱う
                    slash_rel_dir = slash_abs_dir


                res["abs_dir/"          ] = slash_abs_dir
                res["abs_dir/file"      ] = f"{slash_abs_dir}/{file_name}"
                res["abs_dir/file.ext"  ] = f"{slash_abs_dir}/{file_name_ext}"
                res["rel_dir/"          ] = slash_rel_dir
                res["rel_dir/file"      ] = f"{slash_rel_dir}/{file_name}"
                res["rel_dir/file.ext"  ] = f"{slash_rel_dir}/{file_name_ext}"
                res[".ext"              ] = extension
                res["ext"               ] = extension.replace(os.extsep,"")
                res["extension_all"     ] = extension_all
                res["extension"         ] = extension
                res["file"              ] = file_name
                res["file_name"         ] = file_name
                res["file.ext"          ] = file_name_ext
                res["file_label"        ] = file_label

                pass
            case "dir":
                # ディレクトリパスであることが確認されているため
                # ファイル名などは空のままにしておく
                # ディレクトリ関係のパスのみ処理を行う
                slash_abs_dir = slash_abs_path
                slash_rel_dir = os.path.relpath(slash_abs_path)
                # file_name
                # file_name_ext
                # extension
                # extension_all
                res["abs_dir/"          ] = slash_abs_dir
                res["abs_dir/file"      ] = ""
                res["abs_dir/file.ext"  ] = ""
                res["rel_dir/"          ] = os.path.relpath(slash_abs_dir)
                res["rel_dir/file"      ] = ""
                res["rel_dir/file.ext"  ] = ""
                res[".ext"              ] = ""
                res["ext"               ] = ""
                res["extension_all"     ] = ""
                res["extension"         ] = ""
                res["file"              ] = ""
                res["file_name"         ] = ""
                res["file.ext"          ] = ""
                res["file_label"        ] = ""
                pass
            case _:
                raise Exception(
                    f"パスのファイル判定ロジックに\
                    異常が検出されました\
                    {inspect.currentframe().f_code.co_name}\
                    を修正してください。")

        # バックスラッシュパターンも登録する
        bs = dict()
        for k,v in res.items():
            if "/" in k:
                kb = k.replace("/","\\")
                vb = v.replace("/","\\")
                bs[kb]=vb
        res.update(bs)

        res["is_abs"            ] = is_abs
        res["is_exist"          ] = is_exist
        res["is_file"           ] = is_file
        res["is_dir"            ] = is_dir

        res = dict(sorted(res.items()))
        return res
    @staticmethod
    def is_exist(path):
        """ファイルの有無確認
        """
        return os.path.exists(path)
    @staticmethod
    def is_file(path):
        """ファイルかどうかを確認
        """
        return os.path.isfile(path)
    @staticmethod
    def is_dir(path:str):
        """ディレクトリかどうかを判別する
        """
        return os.path.isdir(path)
    @staticmethod
    def get_file_size(path:str):
        """ファイルサイズ取得
        """
        return os.path.getsize(path)
    @staticmethod
    def get_file_update_time(path:str):
        """最終更新日時取得
        """
        return os.path.getmtime(path)
    @staticmethod
    def __globs(mask,recursive=True,sep=os.sep):
        res = list()

        if type(mask) is list:
            for m in mask:
                g = glob.glob(m,recursive=recursive)
                res.extend(g)
        else:
            g = glob.glob(mask,recursive=recursive)
            res.extend(g)
        res = [item.replace("\\",sep).replace("/",sep) for item in set(res)]
        return res
    @staticmethod
    def get_path_list(enable_masks:list[str],disable_masks:list[str]=[],recursive=True)->list:
        enables = ul.__globs(enable_masks,recursive)
        disables = ul.__globs(disable_masks,recursive)
        res = list(set(enables) - set(disables))
        return res
    @staticmethod
    def get_file_sha256(path:str):
        """SHA256取得
        """
        res = ""
        with open(path,"rb") as file:
            fileData = file.read()
            res = hashlib.sha256(fileData).hexdigest()
        return res
    @staticmethod
    def load_bin(path:str):
        res = None
        with open(path, 'rb') as file:
            res = file.read()
        return res
    @staticmethod
    def load_text(path:str,encoding:str=default_encoding)->str:
        res = ""
        with open(path, encoding=encoding) as f:
            res = f.read()
        return res
    @staticmethod
    def load_text_lines(path:str,encoding:str=default_encoding)->list[str]:
        res = []
        with open(path, encoding=encoding) as f:
            res = f.readlines()
        return res
    @staticmethod
    def save_text(path:str,text:str,encoding:str=default_encoding):
        ul.make_file_dir(path)
        with open(path, mode="w", encoding=encoding) as f:
            f.write(text)
    @staticmethod
    def save_text_lines(path:str,lines:list[str],encoding:str=default_encoding):
        ul.make_file_dir(path)
        with open(path, mode="w", encoding=encoding) as f:
            f.writelines(lines)
    @staticmethod
    def load_json(path: str) -> dict:
        text = ul.load_text(path)
        res = json.loads(text)
        return res
    @staticmethod
    def save_json(path:str,data,encoding:str=default_encoding):
        ul.make_file_dir(path)
        with codecs.open(path , mode="w", encoding=encoding) as f:
            json.dump(data, f, ensure_ascii=False,indent=4, sort_keys=True)
    @staticmethod
    def update_json(path: str,data, encoding:str=default_encoding, no_updated_file_over_write=True):
        """ jsonファイルとしてデータを保存する
            既存のデータと内容が同じ場合は書き込みを行いません。
        """
        ea=False
        new_data = data
        old_data_str = ""
        if isinstance(data,dict):
            new_data = dict(sorted(data.items()))
            old_data_str = "{}"
        elif isinstance(data,list):
            new_data = list(sorted(data))
            old_data_str = "[]"
        if no_updated_file_over_write:
            ul.save_json(path,data,encoding)
        else:
            try:
                old_data_str =json.dumps(ul.load_json(path),ensure_ascii=ea,indent=4, sort_keys=True)
                new_data_str =json.dumps(new_data,ensure_ascii=ea,indent=4, sort_keys=True)
            except:
                old_data_str = "{}"
            old_hash = ul.cnv_text_to_sha256(old_data_str)
            new_hash = ul.cnv_text_to_sha256(new_data_str)
            if old_hash == new_hash:
                pass
            else:
                ul.save_json(path,data,encoding)
    @staticmethod
    def load_yaml(path:str, encoding:str=default_encoding) -> dict:
        obj = dict()
        with open(path, encoding=encoding) as f:
            obj = yaml.safe_load(f)
            return obj
    @staticmethod
    def save_yaml(path: str,obj, encoding:str=default_encoding):
        ul.make_file_dir(path)
        with open(path, mode="w", encoding=encoding) as f:
            yaml.dump(obj,f, allow_unicode=True, default_flow_style=False)
            return
    @staticmethod
    def copy_file(src,dest):
        ul.make_file_dir(dest)
        shutil.copy2(src,dest)
    @staticmethod
    def copy_update_files(src_files:list[str],dest_dir:str)->bool:
        res = True
        for src_file in src_files:
            ul.copy_update_file(src_file)
            dest_file_name = ul.cnv_file_name_ext(src_file)
            dest_file_path = os.path.join(dest_dir,dest_file_name)
            res_copy_update_file = ul.copy_update_file(src_file,dest_file_path)
            if not(res_copy_update_file):
                res = False
        return res
    @staticmethod
    def copy_update_file(src:str,dest:str)->bool:
        """ファイルの内容に変化がある場合は更新のため上書きコピーします
        日付の変更のみの場合はコピーしません。
        ファイルパスの指定にはマスクは使えません。
        """
        if not(ul.is_exist(src)):
            raise FileNotFoundError(f'ファイル"{src}"がありません')

        run = False
        if not(ul.is_exist(dest)):
            ul.copy_file(src,dest)
            run = True
        else:
            src_size = ul.get_file_size(src)
            dst_size = ul.get_file_size(dest)

            if not(src_size==dst_size):
                ul.copy_file(src,dest)
                run = True
            else:
                src_sha  = ul.get_file_sha256(src)   
                dst_sha  = ul.get_file_sha256(dest)   
                if not(src_sha==dst_sha):
                    ul.copy_file(src,dest)
                    run = True
                else:
                    # コピーしない(日付のみ変更は考慮外)
                    pass
        return run
    @staticmethod
    def remove_file(path:str):
        """ファイルを削除する
        """
        os.remove(path)
    @staticmethod
    def replace_file(src:str,dest:str):
        """ファイルを置換する(元ファイルは削除される)
        """
        os.replace(src,dest)
    @staticmethod
    def remove_dir(path:str):
        shutil.rmtree(path=path,ignore_errors=True)
    @staticmethod
    def remove_empty_dir(path):
        """指定ディレクトリ以下の空ディレクトリを削除する
        """
        rm = True
        while rm:
            dirs = glob.glob(f'{path}/**/', recursive=True)
            rm = False
            for dir in dirs:
                try:
                    os.removedirs(dir)
                    rm = True
                except OSError:
                    pass
    @staticmethod
    def make_dir(path:str):
        os.makedirs(name=path,exist_ok=True)
    @staticmethod
    def make_file_dir(path:str):
        dir_path = ul.cnv_dir_path(path)
        ul.make_dir(dir_path)
    @staticmethod
    def copy_dir_tree(src:str, dest:str, not_ptn=None):
        """ディレクトリツリーをコピーします
        """
        if not_ptn is None:
            shutil.copytree(src, dest, dirs_exist_ok=True)
        else:
            shutil.copytree(src, dest, ignore=not_ptn, dirs_exist_ok=True)
    #------------------------------------------------------------------------
    # HASH
    #------------------------------------------------------------------------
    @staticmethod
    def cnv_text_to_sha256(src:str)->str:
        return hashlib.sha256(src.encode()).hexdigest()
    @staticmethod
    def cnv_text_to_sha512(src:str)->str:
        return hashlib.sha512(src.encode()).hexdigest()
    @staticmethod
    def cnv_file_to_sha256(path:str)->str:
        fileData = ul.load_bin(path)
        res = hashlib.sha256(fileData).hexdigest()
        return res
    #------------------------------------------------------------------------
    # MARKDOWN
    #------------------------------------------------------------------------
    @staticmethod
    def cnv_markdown_to_html(src:str, include_base_path=".")->str:
        md = markdown.Markdown(extensions=[
            ul.__md_extension(),
            "admonition",
            "tables",
            "toc",
            "nl2br",
            "codehilite",
            "fenced_code",
            MarkdownInclude(configs={'base_path':include_base_path}),
            CustomLinkExtension(),
            ]
            )
        res = md.convert(src)
        return res
    class __md_preprocessor(markdown.preprocessors.Preprocessor):

        has_mermaid = False

        def run_lines(self, lines):
            # ca = code area
            pat_ca_start = re.compile(
                r"^(?P<code_area_sign>[\~\`]{3})[\ \t]*(?P<code_area_name>[a-zA-Z_]*)[\ \t]*$")
            pat_ca_end = re.compile(
                r"^(?P<code_area_sign>[\~\`]{3})[\ \t]*$")

            result_lines = []
            code_area_sign = ""
            code_area_name = ""
            for line in lines:
                code_changed = False
                src_line = line
                print_line = ''.join(filter(lambda x: x in string.printable, line))

                match code_area_name:
                    case "":
                        mat = pat_ca_start.match(print_line)
                        if mat:
                            code_area_sign = mat.group("code_area_sign")
                            code_area_name = mat.group("code_area_name")
                            if code_area_name == "mermaid":
                                result_lines.append('<div class="mermaid">')
                                self.has_mermaid = True
                                code_changed = True
                            else:
                                code_area_name = "-"

                    case "mermaid":
                        mat = pat_ca_end.match(print_line)
                        if mat:
                            if mat.group("code_area_sign") == code_area_sign:
                                result_lines.append("</div>")
                                result_lines.append("")
                                code_area_name = "" # area end
                                code_changed = True
                        if not code_changed:
                            result_lines.append(src_line.strip())
                            code_changed = True

                    case "-": # not supported area
                        mat = pat_ca_end.match(print_line)
                        if mat:
                            if mat.group("code_area_sign") == code_area_sign:
                                code_area_name = "" # area end
                        pass

                if not code_changed:
                    result_lines.append(src_line)

            return result_lines

        def run(self, lines):
            result_lines = self.run_lines(lines)
            if self.has_mermaid:
                result_lines.append("")
                result_lines.append("<script>mermaid.initialize({startOnLoad:true});</script>")
                result_lines.append("")
            return result_lines
    class __md_extension(markdown.Extension):
        def extendMarkdown(self, md, md_globals=None):
            """ add extension instance. """
            md.preprocessors.register(ul.__md_preprocessor(md), 'my_pre_proc', 35)
            md.registerExtension(self)

        def makeExtension(**kwargs):
            return ul.__md_extension(**kwargs)
    #------------------------------------------------------------------------
    # JINJA2
    #------------------------------------------------------------------------
    @staticmethod
    def cnv_template_to_text(data:dict,template_path:str)->str:

        cur = os.getcwd()
        tmp_dir = os.path.dirname(template_path)
        os.chdir(tmp_dir)
        tmp_file = os.path.basename(template_path)

        loader         = jinja2.FileSystemLoader( searchpath=tmp_dir, encoding='utf-8')
        environment    = jinja2.Environment(loader=loader)
        environment.filters["debug"]=ul.log_debug

        temp_data = dict()
        temp_data["4d22a990e03e4ff0b66061daa1674a0d"]=dict(os.environ)
        temp_data["root"]=data

        try:
            template  = environment.get_template(name=tmp_file)
            out_text  = template.render(temp_data)
        except jinja2.exceptions.UndefinedError as e:
            ul.log_debug( f"テンプレートへ入力しているデータをダンプします。")
            ul.log_debug(temp_data)
            ul.log_debug(traceback.format_exc())
            ul.log_debug(e)
            sys.exit(1)
        except jinja2.exceptions.TemplateNotFound as e:
            ul.log_debug(traceback.format_exc())
            ul.log_debug(e)
            ul.log_debug( f"template_path : {template_path}")
            ul.log_debug( f"tmp_dir       : {tmp_dir}")
            ul.log_debug( f"tmp_file      : {tmp_file}")
            ul.log_debug( f"cur           : {cur} -> {os.getcwd()}")
            ul.log_debug( f"raise ファイル「{e}」が見つかりませんでした。")
            sys.exit(1)

        os.chdir(cur)

        return out_text
    #------------------------------------------------------------------------
    # DATETIME
    #------------------------------------------------------------------------
    @staticmethod
    def get_now():
        """現在の日時をdict形式で取得します
        """
        return ul.cnv_datetime_to_dict(datetime.datetime.now())
    @staticmethod
    def cnv_datetime_to_dict(src:datetime):
        res = dict()
        res["YYYYMMDDHHMMSS"] = src.strftime("%Y%m%d%H%M%S")
        res["YYYYMMDD"] = src.strftime("%Y%m%d")
        res["YYYY"] = src.strftime("%Y")
        res["MM"] = src.strftime("%m")
        res["DD"] = src.strftime("%d")
        res["HHMMSS"] = src.strftime("%H%M%S")
        res["YYYY/MM/DD"] = ul.cnv_datetime_to_YYYYMMDD(src,"/")
        res["YYYY-MM-DD"] = ul.cnv_datetime_to_YYYYMMDD(src,"-")
        res["HH:MM:SS"] = src.strftime("%H:%M:%S")
        res["YYYY-MM-DD HH:MM:SS"] = src.strftime("%Y-%m-%d %H:%M:%S")
        res["YYYY年MM月DD日"] = src.strftime("%Y年%m月%d日")
        res["YYYY年MM月DD日(曜日)"] = src.strftime("%Y年%m月%d日(%A)")
        res["YYYY年MM月DD日(曜日short)"] = src.strftime("%Y年%m月%d日(%a)")
        res["曜日"]=src.strftime("%A")
        res["曜日short"]=src.strftime("%a")
        return res
    @staticmethod
    def cnv_datetime_to_YYYYMMDD(src:datetime, sep="/"):
        res = src.strftime(f"%Y{sep}%m{sep}%d")
        return res
    @staticmethod
    def cnv_str_to_datetime(src:str, format:str="%Y/%m/%d %H:%M:%S"):
        """文字列をdatetimeに変換します
        """
        res = datetime.datetime.strptime(src, format)
        return res
    @staticmethod
    def cnv_datetime_to_serial( src ):
        """datetimeをシリアル値に変換します
        """
        zero_day = datetime.datetime(1899,12,31)
        irregular_next = datetime.datetime(1900,3,1)
        res = (src - zero_day).days
        if src >= irregular_next:
            res = res + 1
        return res
    @staticmethod
    def add_days(src,value):
        """datetimeに指定の日数を加算または減算します
        """
        td = datetime.timedelta(days=value)
        return (src + td)
    @staticmethod
    def update_datetime(src:datetime.datetime, year:int=-1,month:int=-1,day:int=-1, hour:int=-1,minute:int=-1,second:int=-1,microsecond:int=-1):
        """日時の一部を更新して返します
        """
        if year == -1:
            year = src.year
        if month == -1:
            month = src.month
        if day == -1:
            day = src.day
        if hour == -1:
            hour = src.hour
        if minute == -1:
            minute = src.minute
        if second == -1:
            second = src.second
        if microsecond == -1:
            microsecond = src.microsecond
        res = datetime.datetime(year=year,month=month,day=day,hour=hour,minute=minute,second=second,microsecond=microsecond)
        return res
    #------------------------------------------------------------------------
    # XLS
    #------------------------------------------------------------------------
    @staticmethod
    def load_xl(path:str)->openpyxl.workbook.workbook.Workbook:
        res = openpyxl.load_workbook(path, keep_vba=False)
        return res
    @staticmethod
    def save_xl(path:str,wb:openpyxl.workbook.workbook.Workbook):
        res = wb.save(path)
        return res
    @staticmethod
    def get_xl_sheet(name:str,workbook:openpyxl.workbook.workbook.Workbook)->openpyxl.worksheet.worksheet.Worksheet:
        res = workbook[name]
        return res
    @staticmethod
    def get_xl_cell_adr(address:str,worksheet:openpyxl.worksheet.worksheet.Worksheet):
        res = worksheet[address]
        return res
    @staticmethod
    def get_xl_cell_pos(row:int,column:int,worksheet:openpyxl.worksheet.worksheet.Worksheet):
        res = None
        if row <= 0:
            pass
        elif column <= 0:
            pass
        else:
            res = worksheet.cell(row=row, column=column)
        return res
    @staticmethod
    def set_xl_cell_value(row:int,column:int,value,ws:openpyxl.worksheet.worksheet.Worksheet):
        if row <= 0:
            pass
        elif column <= 0:
            pass
        else:
            ws.cell(row=row, column=column,value=value)
        return

    #========================================================================
    #
    # FRAMEWORK
    #
    #========================================================================

    #------------------------------------------------------------------------
    # COMMAND LINE APPLICATION
    #------------------------------------------------------------------------
    class cmd_app:

        TYPE_TXT:dict = { "type":"text" }
        TYPE_PATH_NAME:dict = { "type":"path_name" }
        TYPE_PATH_LIST:dict = { "type":"path_list" }

        args_cfg:dict = { "py":TYPE_TXT }
        prefix:str=""
        args:list[str]=[""]
        params:dict = dict()

        path_list:dict = dict()
        env_id:str = ""
        path_names:dict = dict()

        def __init__(self,cfg={"py":TYPE_TXT}) -> None:
            self.args_cfg = cfg

            try:
                self.path_list = ul.load_path()
            except:
                pass

            self.env_id = ul.get_env_id()
            self.path_names = self.path_list.get(self.env_id,{})

        def __display_head(self):
            pf = self.prefix
            ul.log_debug(f"{pf}---------------------------------------------------")
            ul.log_debug(f"{pf}{self.args[0]}")
            ul.log_debug(f"{pf}---------------------------------------------------")
            ul.log_debug(f"{pf}cur_dir  : {os.getcwd()}")
            ul.log_debug(f"{pf}sys.argv : {self.args}")
            ul.log_debug(f"{pf}args_cfg : {self.args_cfg}")

        def __set_params(self):
            pf = self.prefix
            ptn = re.compile("^-(?P<key>[a-zA-Z][0-9a-zA-Z_]*):(?P<value>.*)$")
            arg_list=list(self.args_cfg.keys())
            key_max = max([len(x) for x in arg_list])
            for i in range(0,len(self.args),1):
                k = arg_list[i]
                v = self.args[i]
                m = ptn.match(v)
                if m:
                    k = m.group("key")
                    v = m.group("value")
                    self.params[k]=v
                    ul.log_debug( f"{pf}{k.ljust(key_max)} : {v}")
                else:
                    self.params[k]=v
                    ul.log_debug( f"{pf}{k.ljust(key_max)} : {v}")
            self.params["vscode_debug"]=ul.is_vscode_debug()

        def __set_args_value(self):

            for cfg_key,cfg_val in self.args_cfg.items():
                cfg_type = cfg_val.get("type","")
                prm_val  = self.params[cfg_key]
                match cfg_type:
                    case "text":
                        pass
                    case "list":
                        self.params[cfg_key] = list(prm_val.split(os.pathsep))
                        pass
                    case "path_name":
                        if prm_val in self.path_names.keys():
                            self.params[cfg_key] = self.path_names[prm_val]
                        else:
                            raise Exception(f"(起動引数({cfg_key})で指定されたパス名({prm_val})は存在しません。")
                    case _:
                        pass

            return

        def start(self,name,func,args:list=sys.argv):
            if ul.is_vscode_debug():
                ul.log_enable()
            if name == "__main__":
                self.prefix = f"{name} > "
                self.args = args
                self.__display_head()
                self.__set_params()
                self.__set_args_value()
                func(**self.params)


        @staticmethod
        def __gen_txt(self,py:str, src_path:str,temp_path:str,dest_path:str, vscode_debug:bool):
            if vscode_debug:
                ul.log_enable()

            src_data = {}
            if src_path != "":
                src_data = ul.load_yaml(src_path)

            dest_text = ul.cnv_template_to_text(src_data,temp_path)

            ul.save_text(dest_path,dest_text)
            return

        @staticmethod
        def __gen_txt_name(self,py:str, src_name:str,temp_name:str,dest_name:str, vscode_debug:bool):
            ul.cmd_app.__gen_txt(py,src_name,temp_name,dest_name,vscode_debug)

        @staticmethod
        def start_ex(name:str,app_type:str):
            """定型化されたコマンドラインアプリケーションを開始する。

            Parameters
            ----------
            name : str
                起動元の__name__をセットしてください。
            app_type : str
                アプリケーションのタイプをセットしてください。
                現在指定できるのは下記となります。
                gen_txt(path) : パス指定でYAMLデータファイルとテンプレートファイルからテキストを生成する
                gen_txt(name) : パスリストファイルのパス名でYAMLデータファイルとテンプレートファイルからテキストを生成する

            Raises
            ------
            Exception
                _description_
            """
            app = ul.cmd_app()
            match app_type:
                case "gen_txt(path)":
                    app.args_cfg={
                        "py"        : ul.cmd_app.TYPE_TXT,
                        "src_path"  : ul.cmd_app.TYPE_TXT,
                        "temp_path" : ul.cmd_app.TYPE_TXT,
                        "dest_path" : ul.cmd_app.TYPE_TXT
                    }
                    app.start(name,ul.cmd_app.__gen_txt)
                case "gen_txt(name)":
                    app.args_cfg={
                        "py"        : ul.cmd_app.TYPE_TXT,
                        "src_name"  : {"type":"text"},
                        "temp_name" : {"type":"text"},
                        "dest_name" : {"type":"text"}
                    }
                    app.start(name,ul.cmd_app.__gen_txt_name)
                case _:
                    raise Exception(f"未対応のapp_type({app_type})が指定されました。")

    #------------------------------------------------------------------------
    # PATH CONTROL
    #------------------------------------------------------------------------
    def get_path_list_path()->str:
        return os.environ.get("path_list_path","")
    def get_env_id():
        return os.environ.get("env_id","a48c5c3ad4e94017bcc275492c101193")
    def update_path(src:dict, id:str=get_env_id())->dict:
        dest = ul.load_path()
        src_id = src.get(id,{})
        dest_id = dest.get(id,{})
        dest_id.update(src_id)
        dest[id]=dest_id
        path = ul.get_path_list_path()
        ul.save_json(path,dest)

    def load_path()->dict:
        path = ul.get_path_list_path()
        ul.log_debug(f"{ul.get_caller_function_name()} : path : {path}")
        plp = None
        try:
            plp = ul.load_json(path)
        except:
            plp = dict()
        return plp
